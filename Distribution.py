import openpyxl

Book = openpyxl.load_workbook('<path>')

BookSheet = Book['RAND']
match = {}

for i in range(946):
    if i == 0:
        i += 1
    
    else:
        number = BookSheet.cell(row=i,column=1).value
        seatNumber = BookSheet.cell(row=i,column=2).value
        match[number] = seatNumber
        i += 1

A = 4
stair = 4
WriteSheet = Book['distribution']

for i in match:
    B = A % 3

    if B == 1 :
        WriteSheet.cell(row=stair,column = A).value = i
    elif B == 0 :
        WriteSheet.cell(row=stair,column = A).value = i
    elif B == 2 :
        WriteSheet.cell(row=stair,column = A).value = i
        A -= 15
        stair += 8

    A += 5

C = 4
stair = 5

for i in match.values():
    B = C % 3

    if B == 1 :
        WriteSheet.cell(row=stair,column = C).value = i
    elif B == 0 :
        WriteSheet.cell(row=stair,column = C).value = i
    elif B == 2 :
        WriteSheet.cell(row=stair,column = C).value = i
        C -= 15
        stair += 8

    C += 5




Book.save('<path>')

