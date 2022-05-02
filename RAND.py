import openpyxl
import random


Book = openpyxl.load_workbook('<path>')

BookSheet = Book['RANDMaster']

number = []

for i in range(999999):
    randamNumber = random.randint(100000, 999999)
    if randamNumber in number:
            pass
    else:
            number.append(randamNumber)

    if len(number) == 945:
        break
    else:
        pass

print(number)
print(len(number))

student_number = []
student_number1 = []
student_number2 = []
student_number3 = []

for i in range(8):
    if i == 0:
        i += 1
        pass

    else:
        class_number1 = i * 100 + 1000
        for j in range(46):
            if j ==0 :
                j += 1
                pass
            else:
                StudentNumber1 = class_number1 + j
                student_number1.append(StudentNumber1)
                student_number.append(StudentNumber1)

        class_number2 = i * 100 + 2000
        for k in range(46):
            if k ==0 :
                k += 1
                pass
            else:
                StudentNumber2 = class_number2 + k
                student_number2.append(StudentNumber2)
                student_number.append(StudentNumber2)

        class_number3 = i * 100 + 3000
        for l in range(46):
            if l ==0 :
                l += 1
                pass
            else:
                StudentNumber3 = class_number3 + l
                student_number3.append(StudentNumber3)
                student_number.append(StudentNumber3)

student_number = sorted(student_number)

print(student_number1)
print(student_number2)
print(student_number3)
print(student_number)

for i in range(3):
    for j in range(315):
        sendNumber = student_number1[j]
        randomNumber = number[j]
        j += 1
        BookSheet.cell(row = j , column = 1 ).value = sendNumber
        BookSheet.cell(row=j,column=2).value = randomNumber
    
    for j in range(315):
        sendNumber = student_number2[j]
        randomNumber = number[j + 315]
        j += 1
        BookSheet.cell(row=j,column=4).value = sendNumber
        BookSheet.cell(row=j,column=5).value = randomNumber

    for j in range(315):
        sendNumber = student_number3[j]
        randomNumber = number[j + 630]
        j += 1
        BookSheet.cell(row=j,column=7).value = sendNumber
        BookSheet.cell(row=j,column=8).value = randomNumber


Write_Sheet = Book['RAND']
cellNumber = 1
CellNumber = 1
for i in student_number:
    Write_Sheet.cell(row=cellNumber,column=1).value = i
    cellNumber += 1

for i in number:
    Write_Sheet.cell(row=CellNumber,column=2).value = i
    CellNumber += 1




Book.save('<path>')
