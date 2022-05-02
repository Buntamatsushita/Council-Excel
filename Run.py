import openpyxl

Book = openpyxl.load_workbook('<path>')

BookSheet = Book['result']
number = []
SheetCharactor = []

for i in range(451):
    if i == 0 or i == 1 or i == 2:
        i += 1
    else:
        AA = BookSheet.cell(row=i,column=2).value
        if AA != None:
            SheetCharactor.append(str(AA))
        i += 1

for i in range(453):
    if i == 0 or i == 1 or i == 2:
        i += 1
    else:
        AA = BookSheet.cell(row=i,column=3).value
        if AA != None:
            number.append(str(AA))
        else:
            pass
        i += 1

WriteSheet = Book['Sheet']

wa = 1
for i in number:
    WriteSheet.cell(row=wa,column=1).value = int(i)
    WriteSheet.cell(row=wa,column=2).value = "あなたは第１部に当選しました。"
    wa += 1

wb = 1
for i in SheetCharactor:
    WriteSheet.cell(row=wb,column=3).value = "座席は" + i + "ブロックです。"
    wb += 1


Book.save('<path>')