import openpyxl

Book = openpyxl.load_workbook('<path>')

BookSheet = Book['フォームの回答']
part1 = []
part2 = []

for i in range(850):
    if i == 0 or i == 2:
        i += 1
    else:
        number = BookSheet.cell(row=i,column=2).value
        part = BookSheet.cell(row=i,column=3).value
        if number != None:
            if part == "１部":
                part1.append(str(number))
            else:
                part2.append(str(number))
        i += 1


Book2 = openpyxl.load_workbook('<path>')
BookSheet2 = Book['Match']
match = {}

for i in range(946):
    if i == 0:
        i += 1
    
    else:
        number = BookSheet2.cell(row=i,column=1).value
        seatNumber = BookSheet2.cell(row=i,column=2).value
        match[number] = seatNumber
        i += 1

numberPart1 = []

for i in part1:
    if i in match.keys():
        send = match[i]
        numberPart1.append(send)

numberPart2 = []

for i in part2:
    if i in match.keys():
        send = match[i]
        numberPart2.append(send)





Book2 = openpyxl.load_workbook('<path>')
WriteSheet = Book2['result']

wa = 3
for i in numberPart1:
    WriteSheet.cell(row=wa,column=3).value = int(i)
    wa += 1

wa = 3
for i in numberPart2:
    WriteSheet.cell(row=wa,column=9).value = int(i)
    wa += 1

Book.save('<path>')